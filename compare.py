import matlab
import argparse
import pandas as pd
import time
import subprocess
from datetime import datetime
import yaml
from subprocess import *
from rosbag.bag import Bag
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.merge import MergedCell
from datetime import datetime 
import src.trajectory as tj
import src.time as tc

import src.error as error

packages_list = {
    'aloam': ['record_aloam.launch', 'record_aloam_nclt.launch'],
    'lego_loam': ['record_lego_loam.launch', 'record_lego_loam_nclt.launch'],
    'lio_sam': ['record_lio_sam.launch', 'record_lio_sam_nclt.launch'],
    'kiss_icp': ['record_kiss_icp.launch', 'record_kiss_icp_nclt.launch'],
    'dlo': ['record_dlo.launch', 'record_dlo_nclt.launch'],
    'f_loam': ['record_f_loam.launch', 'record_f_loam_nclt.launch'],
    'faster_lio': ['record_faster_lio.launch', 'record_faster_lio_nclt.launch'],
    'fast_lio': ['record_fast_lio.launch', 'record_fast_lio_nclt.launch']
}

parser = argparse.ArgumentParser(description="Lidar SLAM Evaluator")
parser.add_argument('--dataset', nargs="+", dest='dataset', help='SLAM algorithms that want to compare: kitti, nclt')
parser.add_argument('--slam', nargs="+", dest='slam_packages', help='SLAM algorithms that want to compare: aloam, lego_loam, lio_sam')
parser.add_argument('--bag_path', dest='bagfile_path', help='Directory path where KITTI dataset bag file "kitti.bag" exists')
parser.add_argument('--convert_kitti', dest='convert_kitti_arg', help='If you want to convert bag to kitti txt file, add --convert_kitti <seq number>')
parser.add_argument('--plot', dest='plot_arg', choices=['all', 'traj', 'error', 'stat'], default='all', help='Plot options: all, traj, error, stat')
parser.add_argument('--no_play', dest='play_flag', action='store_false', help='If you already have recorded result bag file, add --no_play')
args = parser.parse_args()

if args.dataset is None:
    print("No dataset specified. Using KITTI dataset by default.")
    args.dataset = ["kitti"]
if args.slam_packages is None:
    print("No specified SLAM Lists. Run default algorithms: ALOAM, LEGO_LOAM, LIO_SAM")
    args.slam_packages = ["aloam", "lego_loam", "lio_sam", "kiss_icp"]

if 'all' in args.slam_packages:
    args.slam_packages = ["faster_lio", "fast_lio", "dlo", "kiss_icp","f_loam"]

for slam in args.slam_packages:
    if slam not in packages_list:
        raise ValueError("%s algorithm is not available!" % slam)

if args.bagfile_path is None:
    raise parser.error("Please set bag file directory with --bag_path parser")
# if args.convert_kitti_arg is None:
#    raise parser.error("Please set seq number --convert_kitti_arg 07")

class CompareSLAM():
    def __init__(self, slam_packages, bag_path, plot_arg, convert_kitti_arg):
        self.slam_packages = slam_packages
        self.bag_path = bag_path
        self.bag_file = bag_path + '/' + args.dataset[0] + '.bag'
        bag_info_dict = yaml.safe_load(Bag(self.bag_file, 'r')._get_yaml_info())
        self.bag_duration = bag_info_dict['duration']
        self.plot_arg = plot_arg
        self.convert_kitti_arg = convert_kitti_arg
        self.file_list = []
        self.file_list_time = []
        print(args.dataset)
        self.file_list.append(bag_path + '/' + args.dataset[0] + '_gt.bag')
        for slam in self.slam_packages:
            self.file_list.append(self.bag_path + '/' + slam + '_path.bag')
            self.file_list_time.append(self.bag_path + '/' + slam + '_time.bag')

    def play_algorithm(self):     
        for slam in self.slam_packages:            
            print("%s algorithm is running ..." % slam)
            print("bag file duration: %i" % self.bag_duration)
            print(self.bag_file)
            
            p1 = Popen(["roslaunch", "path_recorder", 'record_' + slam + '_' + args.dataset[0] + '.launch', "bag_path:=" + self.bag_path])
            
            start_time = time.time()
            current_time = start_time
            first = True

            while True:
                time.sleep(1)
                if first and args.dataset[0] == "ntu":
                    odom_t = "/robot/dlo/odom_node/odom" if slam == "dlo" else "/Odometry"
                    
                    # Ejecutar ambos comandos en paralelo
                    comp_time_cmd = f"timeout {str(self.bag_duration)} rostopic echo -p --nostr --noarr /comp_time > {str(self.bag_path)}/result/{slam}_comp_time.csv"
                    odom_cmd = f"timeout {str(self.bag_duration)} rostopic echo -p --nostr --noarr {odom_t} > {str(self.bag_path)}/result/{slam}_predict_odom.csv"

                    comp_time_process = Popen(comp_time_cmd, shell=True)
                    odom_process = Popen(odom_cmd, shell=True)

                    # Esperar a que ambos procesos terminen
                    comp_time_process.wait()
                    odom_process.wait()
                    
                    first = False
                
                current_time = time.time()
                if current_time - start_time >= self.bag_duration + 15:
                    print("finishing %s ..." % slam)
                    p1.terminate()
                    p1.wait()
                    break
            
        print("Finished all algorithms")
        self.plot()

    def plot(self):
        if args.dataset[0] == "ntu":
            eng = matlab.engine.start_matlab()
            eng.run("/home/dronomy/TFM_ws/NTU_data/viral_eval/evaluate_all.m", nargout=0)
            return
        if self.convert_kitti_arg and args.dataset[0] == "kitti":
            for slam in self.slam_packages:
                print("Converting %s algorithm bag to KITTI dataset format..." % slam)
                p2 = Popen(["python3", "bag2kittiformat2.py", slam, self.convert_kitti_arg])
                p2.wait()
        plot_arg = self.plot_arg
        file_list = self.file_list
        gt, tj_list = self.traj_process(file_list)
        time_list = self.time_process(self.file_list_time)
        if plot_arg == 'traj':
            self.plot_traj(gt, tj_list)
        else:
            error_list = self.error_process(gt, tj_list)
            self.plot_error(plot_arg, gt, tj_list, error_list, time_list)
            plt.close('all')  # Cerrar todas las figuras abiertas

    def time_process(self, data_files):
        time_list = []
        for file in data_files:
            if file.endswith('.bag') or file.endswith('.txt'):
                trajectory = tc.TimeComp(file)
                time_list.append(trajectory)
            else:
                print("Unsupported .{} file type".format(file.split('.')[-1]))
        return time_list

    def traj_process(self, data_files):
        tj_list = []
        gt = None
        for file in data_files:
            if file.endswith('.bag') or file.endswith('.txt'):
                trajectory = tj.Trajectory(file)
                if not trajectory.is_gt:
                    tj_list.append(trajectory)
                else:
                    gt = trajectory
            else:
                print("Unsupported .{} file type".format(file.split('.')[-1]))
        return gt, tj_list

    def error_process(self, gt, tj_list):
        error_list = []
        if gt is None:
            print('Need ground truth for error calculation.')
            return
        for tj in tj_list:
            error_list.append(error.Error(gt, tj))
        return error_list

    def plot_traj(self, gt, tj_list):
        print("plotting...")
        tj.plotXYZ(gt, tj_list)
        tj.plot3D(gt, tj_list)
        plt.show(block=False)  # No bloquear la ejecución
        plt.pause(1)  # Pausar para mostrar las figuras
        plt.close('all')  # Cerrar todas las figuras abiertas

    def plot_error(self, plot_arg, gt, tj_list, error_list, comp_list):
        print("plotting...")
        if plot_arg == 'all':
            print("tj_list: ", tj_list)
            print("error_list: ", error_list)
            
            tj.plotXYZ(gt, tj_list)
            tj.plotXYZT(gt, error_list, comp_list)
            tj.plot2D('xy', gt, error_list)
            tj.plot3D(gt, tj_list)
            error.plotAPE(error_list)
            error.plotAPEStats(error_list)
            error.plotRPE(error_list)
            error.plotRPEStats(error_list)
            error.printSequenceErrorStats(error_list)
        if plot_arg == 'error':
            error.plotAPE(error_list)
            error.plotAPEStats(error_list)
            error.plotRPE(error_list)
            error.plotRPEStats(error_list)
            error.printSequenceErrorStats(error_list)
        if plot_arg == 'stat':
            error.plotAPEStats(error_list)
            error.plotRPEStats(error_list)
            error.printSequenceErrorStats(error_list)
        self.print_result(error_list, comp_list)
        import os
        # Guardar la gráfica 2D en formato PNG full HD 
        plot_path = os.path.join(self.bag_path, '2D_plot.png')
        tj.plot2D('xy', gt, error_list)
        plt.savefig(plot_path)
        print(f"Gráfica 2D guardada en {plot_path}")
        # Guardar la gráfica XYZT en formato PNG
        plot_path = os.path.join(self.bag_path, 'XYZT_plot.png')
        tj.plotXYZT(gt, error_list, comp_list)
        plt.savefig(plot_path)
        print(f"Gráfica XYZT guardada en {plot_path}")
        plt.show(block=False)  # No bloquear la ejecución
        plt.pause(1)  # Pausar para mostrar las figuras
        plt.close('all')  # Cerrar todas las figuras abiertas

        
    def print_result(self, error_list, comp_list):
        wb = Workbook()
        ws = wb.active

        # Añadir el título de APE[m] con celdas combinadas
        ws.append(['name', 'kitti_t[m]', '', '', '', '', '', 'kitti_rot[rad]', '', '', '', '', '', 'APE[m]', '', '', '', '', '', 'RPE[m]', '', '', '', '', '', 'RPE[rad]', '', '', '', '', '', 'ave_comp[ms]', 'seq length[s]'])
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
        ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=19)
        ws.merge_cells(start_row=1, start_column=20, end_row=1, end_column=25)
        ws.merge_cells(start_row=1, start_column=26, end_row=1, end_column=31)

        # Añadir los encabezados para los datos de APE y RPE
        headers = ['mean', 'std', 'median', 'minimum', 'maximum', 'rmse']
        ws.append([' '] + headers + headers + headers + headers + headers)

        for error, comp_time in zip(error_list, comp_list):
    
            print(error.name)
            print("APE[m]   : {}".format(error.ape_tans_stat[5]))
            print("RPE[m]   : {}".format(error.rpe_tans_stat[5]))
            print("RPE kitti[m]   : {}".format(error.seq_trans_stat[5]))
            print("RPE[rad] : {}".format(error.rpe_rot_stat[0]))
            # data.append({
            # 'Name': error.name,
            # 'APE[m]': error.ape_tans_stat[5],
            # 'RPE[m]': error.rpe_tans_stat[5],
            # 'RPE[rad]': error.rpe_rot_stat[0]
            # })
            comp_time.comp_time_list.mean
            ws.append([error.name] + error.seq_trans_stat[:] + error.seq_rot_stat[:] + error.ape_tans_stat[:] + error.rpe_tans_stat[:] + error.rpe_rot_stat[:] + [comp_time.comp_time_list.mean()] + [self.bag_duration])
        
        # Puedes hacer algo similar para RPE si lo necesitas
        from openpyxl.styles import Alignment   
        # Ajustar el ancho de las columnas
        for column_cells in ws.columns:
            if isinstance(column_cells[0], MergedCell):  # Saltar si es una celda combinada
                continue
            length = max(len(str(cell.value)) for cell in column_cells if not isinstance(cell, MergedCell))
            ws.column_dimensions[column_cells[0].column_letter].width = length

        center_aligned_text = Alignment(horizontal='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_aligned_text

        fecha_hora_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        nombre_archivo = f"{self.bag_path}/resultados_{fecha_hora_actual}.xlsx"
        wb.save(nombre_archivo)
       
if __name__ == '__main__':
    compare = CompareSLAM(args.slam_packages, args.bagfile_path, args.plot_arg, args.convert_kitti_arg)
    if args.play_flag:
        compare.play_algorithm()
    else:
        compare.plot()
